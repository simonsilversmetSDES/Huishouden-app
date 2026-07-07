"""Eenmalige migratie (spec §10) van de vermogenshistoriek uit het Excel-werkboek.

Twee tabbladen:
- **Rekeningstatus** (§6): per context een blok met een `Datum`-koprij; de kolommen
  tussen `Datum` en `Totaal` zijn rekeningen (Zichtrekening, Spaarrekening, …).
  Elke maandrij → `AccountSnapshot` per rekening. Kolommen die niet op een rekening
  van die context mappen (bv. 'Vrije ruimte …', 'Groepsverzekering') worden
  gerapporteerd en overgeslagen.
- **Status balans** (§9): per context een blok; de rij ná de contextnaam bevat de
  maanddatums, de rijen daaronder de activaklassen (→ `AssetClass`) tot `Totaal`.
  Elke niet-nulcel → `NetWorthSnapshot`.

Geld als Decimal op de cent (nooit float). Idempotent via de bestaande UNIQUE-
constraints (opnieuw importeren werkt bestaande standen bij). Draai op een KOPIE
van de database (CLAUDE.md).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Account, AccountSnapshot, Context, NetWorthSnapshot
from app.models.enums import AccountType, AssetClass

CENT = Decimal("0.01")
CONTEXT_NAMES = ("Gemeenschappelijk", "Simon", "Jozefien")

REKENINGSTATUS_SHEET = "Rekeningstatus"
STATUS_BALANS_SHEET = "Status balans"

# Rekeningkolomkop → rekeningtype (voor de mapping op accounts.type).
_TYPE_KEYWORDS: list[tuple[str, AccountType]] = [
    ("zicht", AccountType.ZICHT),
    ("spaar", AccountType.SPAAR),
    ("belegg", AccountType.BELEGGING),
]

# Activaklasse-rijlabel → enum.
_ASSET_KEYWORDS: list[tuple[str, AssetClass]] = [
    ("contant", AssetClass.CONTANT),
    ("belegg", AssetClass.ETF_FONDSEN),
    ("pensioen", AssetClass.PENSIOENSPAREN),
    ("groepsverzeker", AssetClass.GROEPSVERZEKERING),
    ("woning", AssetClass.WONING),
    ("aandelen", AssetClass.AANDELEN),
]


@dataclass
class AccountBlockReport:
    context: str
    snapshots_new: int = 0
    snapshots_updated: int = 0
    mapped_columns: list[str] = field(default_factory=list)  # "kop → rekeningnaam"
    unmapped_columns: list[str] = field(default_factory=list)


@dataclass
class NetWorthBlockReport:
    context: str
    values_new: int = 0
    values_updated: int = 0
    unmapped_rows: list[str] = field(default_factory=list)


@dataclass
class BalanceImportReport:
    accounts: list[AccountBlockReport] = field(default_factory=list)
    net_worth: list[NetWorthBlockReport] = field(default_factory=list)


def _norm(value: Any) -> str:
    return " ".join(str(value).split()).casefold() if isinstance(value, str) else ""


def _to_money(value: Any) -> Decimal:
    return Decimal(str(value)).quantize(CENT, rounding=ROUND_HALF_UP)


def _is_number(value: Any) -> bool:
    return isinstance(value, int | float) and not isinstance(value, bool)


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    return value if isinstance(value, date) else None


def _context_in_row(row: tuple[Any, ...]) -> str | None:
    for value in row:
        if isinstance(value, str) and value.strip() in CONTEXT_NAMES:
            return value.strip()
    return None


class _AccountResolver:
    """Mapt een rekeningkolomkop op een rekening van de juiste context (op type)."""

    def __init__(self, db: Session) -> None:
        self._by_context: dict[int, list[Account]] = {}
        for account in db.scalars(select(Account)):
            self._by_context.setdefault(account.context_id, []).append(account)

    def resolve(self, context: Context, header: str) -> Account | None:
        accounts = self._by_context.get(context.id, [])
        norm = _norm(header)
        wanted = next((t for kw, t in _TYPE_KEYWORDS if kw in norm), None)
        if wanted is None:
            return None
        candidates = [a for a in accounts if a.type == wanted and a.active]
        if len(candidates) == 1:
            return candidates[0]
        # meerdere van hetzelfde type: kies op naamovereenkomst
        for account in candidates:
            if norm in _norm(account.name) or _norm(account.name) in norm:
                return account
        return candidates[0] if candidates else None


def _asset_for(label: str) -> AssetClass | None:
    norm = _norm(label)
    return next((asset for kw, asset in _ASSET_KEYWORDS if kw in norm), None)


def import_rekeningstatus(
    db: Session, ws: Worksheet, contexts: dict[str, Context], resolver: _AccountResolver
) -> list[AccountBlockReport]:
    rows = list(ws.iter_rows(values_only=True))
    existing = {
        (s.account_id, s.snapshot_date): s for s in db.scalars(select(AccountSnapshot)).all()
    }
    reports: list[AccountBlockReport] = []

    context: Context | None = None
    report: AccountBlockReport | None = None
    date_col: int | None = None
    col_account: dict[int, Account] = {}

    for row in rows:
        name = _context_in_row(row)
        if name is not None:
            context = contexts[name]
            report = AccountBlockReport(context=name)
            reports.append(report)
            date_col = None
            col_account = {}
            continue
        if context is None or report is None:
            continue

        # Koprij met "Datum": kolommen tot "Totaal" zijn rekeningen.
        if date_col is None:
            datum_idx = next(
                (i for i, v in enumerate(row) if isinstance(v, str) and _norm(v) == "datum"), None
            )
            if datum_idx is None:
                continue
            date_col = datum_idx
            for i in range(datum_idx + 1, len(row)):
                header = row[i]
                if not isinstance(header, str) or not header.strip():
                    continue
                if _norm(header) in ("totaal", "verandering"):
                    if _norm(header) == "totaal":
                        break
                    continue
                account = resolver.resolve(context, header)
                if account is None:
                    report.unmapped_columns.append(header.strip())
                else:
                    col_account[i] = account
                    report.mapped_columns.append(f"{header.strip()} → {account.name}")
            continue

        # Maandrij: datum in date_col, saldo per rekeningkolom (incl. 0; None = blanco).
        snapshot_date = _as_date(row[date_col]) if date_col < len(row) else None
        if snapshot_date is None:
            continue
        for col, account in col_account.items():
            value = row[col] if col < len(row) else None
            if not _is_number(value):
                continue
            balance = _to_money(value)
            key = (account.id, snapshot_date)
            if key in existing:
                if existing[key].balance != balance:
                    existing[key].balance = balance
                    report.snapshots_updated += 1
            else:
                snap = AccountSnapshot(
                    account_id=account.id, snapshot_date=snapshot_date, balance=balance
                )
                db.add(snap)
                existing[key] = snap
                report.snapshots_new += 1

    return reports


def import_status_balans(
    db: Session, ws: Worksheet, contexts: dict[str, Context]
) -> list[NetWorthBlockReport]:
    rows = list(ws.iter_rows(values_only=True))
    existing = {
        (s.context_id, s.snapshot_date, s.asset_class): s
        for s in db.scalars(select(NetWorthSnapshot)).all()
    }
    reports: list[NetWorthBlockReport] = []
    seen: set[str] = set()  # enkel het eerste blok per context; lager staan aggregaties

    i = 0
    while i < len(rows):
        name = _context_in_row(rows[i])
        if name is None or name in seen:
            i += 1
            continue
        seen.add(name)
        context = contexts[name]
        report = NetWorthBlockReport(context=name)
        reports.append(report)

        # De rij ná de contextnaam bevat de maanddatums per kolom.
        date_row = rows[i + 1] if i + 1 < len(rows) else ()
        col_date = {j: d for j, v in enumerate(date_row) if (d := _as_date(v)) is not None}

        j = i + 2
        while j < len(rows):
            row = rows[j]
            if _context_in_row(row) is not None:
                break
            label = next((v for v in row if isinstance(v, str) and v.strip()), None)
            if label is not None and _norm(label) == "totaal":
                break
            asset = _asset_for(label) if label else None
            if label is not None and asset is None and _norm(label) not in ("",):
                report.unmapped_rows.append(label.strip())
            if asset is not None:
                for col, snapshot_date in col_date.items():
                    value = row[col] if col < len(row) else None
                    if not _is_number(value) or value == 0:
                        continue
                    amount = _to_money(value)
                    key = (context.id, snapshot_date, asset)
                    if key in existing:
                        if existing[key].value != amount:
                            existing[key].value = amount
                            report.values_updated += 1
                    else:
                        snap = NetWorthSnapshot(
                            context_id=context.id,
                            snapshot_date=snapshot_date,
                            asset_class=asset,
                            value=amount,
                        )
                        db.add(snap)
                        existing[key] = snap
                        report.values_new += 1
            j += 1
        i = j

    return reports


def import_balance_workbook(db: Session, path: Path) -> BalanceImportReport:
    """Volledige balans-import; commit pas als beide tabbladen zonder fout verwerkt zijn."""
    report = BalanceImportReport()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        contexts = {c.name: c for c in db.scalars(select(Context))}
        missing = [name for name in CONTEXT_NAMES if name not in contexts]
        if missing:
            raise ValueError(f"Contexten ontbreken in de database (seed eerst): {missing}")

        resolver = _AccountResolver(db)
        if REKENINGSTATUS_SHEET in workbook.sheetnames:
            report.accounts = import_rekeningstatus(
                db, workbook[REKENINGSTATUS_SHEET], contexts, resolver
            )
        if STATUS_BALANS_SHEET in workbook.sheetnames:
            report.net_worth = import_status_balans(
                db, workbook[STATUS_BALANS_SHEET], contexts
            )
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        workbook.close()
    return report
