"""Eenmalige Excel-import (spec §10): Budget Planning + Tracking → database.

Afspraken met Simon (05/07/2026):
- Tekens: bedragen staan positief in de Excel, het teken zit in de Type-kolom.
  App-conventie: Inkomen → +, Uitgaven/Sparen → −; negatieve Excel-bedragen
  (correcties, opnames) keren daardoor vanzelf om.
- Negatieve spaarbudgetten (geplande opnames) blijven negatief budget.
- "Enter … Category"-placeholders en "Total"-rijen worden overgeslagen;
  categorienamen worden op de bestaande lijst gemapt, onbekende categorieën
  worden per context aangemaakt.
- effective_date komt uit de kolom "Effective Date" (fallback: Date).

Idempotent: budgetcellen zijn upserts; transacties krijgen een import_hash op
basis van tabblad + rijnummer, dus hetzelfde werkboek twee keer importeren
voegt niets toe. Het script draait op een KOPIE van de database (CLAUDE.md).
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Budget, Category, Context, Transaction
from app.models.enums import Categorization, CategoryType, TransactionSource

CENT = Decimal("0.01")

BUDGET_SHEET = "Budget Planning"
TRACKING_SHEETS = {
    "Tracking Gem.": "Gemeenschappelijk",
    "Tracking S.": "Simon",
    "Tracking J.": "Jozefien",
}

# Sectiekoppen heten per blok "Inkomen", "Inkomen2", "Inkomen3", …
_SECTION_RE = re.compile(r"^(Inkomen|Uitgaven|Sparen)\d*$")
_BLOCK_PREFIX = "Budget Planning "


@dataclass
class BudgetBlockReport:
    context: str
    cells_new: int = 0
    cells_updated: int = 0
    placeholders_skipped: int = 0
    years: list[int] = field(default_factory=list)


@dataclass
class TrackingReport:
    sheet: str
    context: str
    imported: int = 0
    duplicates: int = 0
    skipped: list[str] = field(default_factory=list)


@dataclass
class ImportReport:
    budget_blocks: list[BudgetBlockReport] = field(default_factory=list)
    tracking: list[TrackingReport] = field(default_factory=list)
    categories_created: list[str] = field(default_factory=list)  # "Context / Type / Naam"
    name_mappings: list[str] = field(default_factory=list)  # "Excelnaam → categorienaam"


class _CategoryResolver:
    """Mapt Excel-namen op bestaande categorieën; maakt onbekende aan per context."""

    def __init__(self, db: Session, report: ImportReport) -> None:
        self.db = db
        self.report = report
        self._cache: dict[tuple[int, CategoryType, str], Category] = {}
        for category in db.scalars(select(Category)):
            self._cache[(category.context_id, category.type, _norm(category.name))] = category

    def resolve(self, context: Context, cat_type: CategoryType, raw_name: str) -> Category:
        name = " ".join(raw_name.split())
        key = (context.id, cat_type, _norm(name))
        if key in self._cache:
            return self._cache[key]

        # Fuzzy: Excel-naam en seed-naam verschillen soms in een suffix
        # (bv. "Ontspanning/Sport/Boeken/…" vs. "Ontspanning/Sport/Boeken").
        for (ctx_id, existing_type, existing_norm), category in self._cache.items():
            if ctx_id != context.id or existing_type != cat_type:
                continue
            shortest = min(len(existing_norm), len(_norm(name)))
            if shortest >= 8 and (
                existing_norm.startswith(_norm(name)) or _norm(name).startswith(existing_norm)
            ):
                self._cache[key] = category
                self.report.name_mappings.append(f"{context.name}: '{name}' → '{category.name}'")
                return category

        max_sort = max(
            (c.sort_order for c in self._cache.values() if c.context_id == context.id),
            default=0,
        )
        category = Category(
            context_id=context.id, name=name, type=cat_type, sort_order=max_sort + 1
        )
        self.db.add(category)
        self.db.flush()
        self._cache[key] = category
        self.report.categories_created.append(f"{context.name} / {cat_type} / {name}")
        return category


def _norm(name: str) -> str:
    return " ".join(name.split()).casefold()


def _to_money(value: Any) -> Decimal:
    """Excel-getal (float/int) → Decimal op de cent, HALF_UP zoals Excel toont."""
    return Decimal(str(value)).quantize(CENT, rounding=ROUND_HALF_UP)


def _is_number(value: Any) -> bool:
    return isinstance(value, int | float) and not isinstance(value, bool)


def _first_text(row: tuple[Any, ...], start: int = 1, stop: int = 4) -> str | None:
    for value in row[start:stop]:
        if isinstance(value, str) and value.strip():
            return value.strip()
    return None


def _is_placeholder(label: str) -> bool:
    lowered = label.casefold()
    return (
        (lowered.startswith("enter") and "categor" in lowered)
        or lowered == "total"
        or lowered.startswith("to be allocated")
        or lowered.startswith("define starting")
    )


def import_budget_sheet(
    db: Session, ws: Worksheet, contexts: dict[str, Context], resolver: _CategoryResolver
) -> list[BudgetBlockReport]:
    """Drie blokken (Gem./Simon/Jozefien) onder elkaar; per sectie een datumkoprij."""
    reports: list[BudgetBlockReport] = []
    current: BudgetBlockReport | None = None
    context: Context | None = None
    section: CategoryType | None = None
    month_cols: dict[int, tuple[int, int]] = {}

    existing = {
        (b.category_id, b.year, b.month): b
        for b in db.scalars(select(Budget)).all()
    }

    for row in ws.iter_rows(values_only=True):
        label = _first_text(row)
        if label is None:
            continue

        if label.startswith(_BLOCK_PREFIX):
            context_name = label.removeprefix(_BLOCK_PREFIX).strip()
            if context_name not in contexts:
                raise ValueError(f"Onbekende context in werkboek: '{context_name}'")
            context = contexts[context_name]
            current = BudgetBlockReport(context=context_name)
            reports.append(current)
            section = None
            month_cols = {}
            continue

        section_match = _SECTION_RE.match(label)
        if section_match:
            section = CategoryType(section_match.group(1))
            month_cols = {
                i: (value.year, value.month)
                for i, value in enumerate(row)
                if isinstance(value, datetime | date)
            }
            continue

        if context is None or current is None or section is None or not month_cols:
            continue
        if _is_placeholder(label):
            current.placeholders_skipped += 1
            continue

        category: Category | None = None  # lazy: geen categorie aanmaken voor lege rijen
        for col, (year, month) in month_cols.items():
            value = row[col] if col < len(row) else None
            if not _is_number(value) or value == 0:
                continue
            if category is None:
                category = resolver.resolve(context, section, label)
            amount = _to_money(value)
            key = (category.id, year, month)
            if key in existing:
                if existing[key].amount != amount:
                    existing[key].amount = amount
                    current.cells_updated += 1
            else:
                budget = Budget(category_id=category.id, year=year, month=month, amount=amount)
                db.add(budget)
                existing[key] = budget
                current.cells_new += 1
            if year not in current.years:
                current.years.append(year)

    for report in reports:
        report.years.sort()
    return reports


def _header_columns(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]] | None:
    """Zoek de koprij (Date + Amount) en geef kolomindexen per kopnaam terug."""
    for i, row in enumerate(rows[:30]):
        headers = {
            str(v).strip().casefold(): j for j, v in enumerate(row) if isinstance(v, str)
        }
        if "date" in headers and "amount" in headers:
            return i, headers
    return None


def _import_hash(sheet: str, row_index: int) -> str:
    return hashlib.sha256(f"excel:{sheet}:{row_index}".encode()).hexdigest()


def import_tracking_sheet(
    db: Session, ws: Worksheet, context: Context, resolver: _CategoryResolver
) -> TrackingReport:
    report = TrackingReport(sheet=ws.title, context=context.name)
    rows = [row for row in ws.iter_rows(values_only=True)]

    found = _header_columns(rows)
    if found is None:
        report.skipped.append("geen koprij met Date + Amount gevonden")
        return report
    header_idx, headers = found

    # S.- en J.-bladen: kolom "Type2" bevat het échte type, "Type" een hulpjeslabel.
    type_col = headers.get("type2", headers.get("type"))
    date_col = headers["date"]
    amount_col = headers["amount"]
    category_col = headers.get("category")
    details_col = headers.get("details")
    effective_col = headers.get("effective date")

    known_hashes = set(
        db.scalars(select(Transaction.import_hash).where(Transaction.import_hash.is_not(None)))
    )

    def cell(row: tuple[Any, ...], col: int | None) -> Any:
        return row[col] if col is not None and col < len(row) else None

    for row_index, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        tx_date = cell(row, date_col)
        amount = cell(row, amount_col)
        if tx_date is None and amount is None:
            continue  # lege rij
        if not isinstance(tx_date, datetime | date):
            report.skipped.append(f"rij {row_index}: geen geldige datum ({tx_date!r})")
            continue
        if not _is_number(amount):
            report.skipped.append(f"rij {row_index}: geen geldig bedrag ({amount!r})")
            continue

        raw_type = cell(row, type_col)
        type_text = str(raw_type).strip() if raw_type is not None else ""
        try:
            tx_type = CategoryType(type_text)
        except ValueError:
            report.skipped.append(f"rij {row_index}: onbekend type '{type_text}'")
            continue

        import_hash = _import_hash(ws.title, row_index)
        if import_hash in known_hashes:
            report.duplicates += 1
            continue
        known_hashes.add(import_hash)

        magnitude = _to_money(amount)
        signed = magnitude if tx_type == CategoryType.INKOMEN else -magnitude

        raw_category = cell(row, category_col)
        category = (
            resolver.resolve(context, tx_type, str(raw_category).strip())
            if isinstance(raw_category, str) and raw_category.strip()
            else None
        )

        effective = cell(row, effective_col)
        if not isinstance(effective, datetime | date):
            effective = tx_date
        details = cell(row, details_col)

        db.add(
            Transaction(
                context_id=context.id,
                category_id=category.id if category else None,
                date=tx_date.date() if isinstance(tx_date, datetime) else tx_date,
                effective_date=(
                    effective.date() if isinstance(effective, datetime) else effective
                ),
                amount=signed,
                type=tx_type,
                description=str(details).strip() if details is not None else None,
                source=TransactionSource.IMPORT_EXCEL,
                import_hash=import_hash,
                categorization=(
                    Categorization.MANUAL if category else Categorization.UNCATEGORIZED
                ),
            )
        )
        report.imported += 1

    return report


def import_workbook(db: Session, path: Path) -> ImportReport:
    """Volledige import; commit pas als alle tabbladen zonder fout verwerkt zijn."""
    report = ImportReport()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        contexts = {c.name: c for c in db.scalars(select(Context))}
        missing = [name for name in TRACKING_SHEETS.values() if name not in contexts]
        if missing:
            raise ValueError(f"Contexten ontbreken in de database (seed eerst): {missing}")

        resolver = _CategoryResolver(db, report)
        if BUDGET_SHEET in workbook.sheetnames:
            report.budget_blocks = import_budget_sheet(
                db, workbook[BUDGET_SHEET], contexts, resolver
            )
        for sheet_name, context_name in TRACKING_SHEETS.items():
            if sheet_name in workbook.sheetnames:
                report.tracking.append(
                    import_tracking_sheet(
                        db, workbook[sheet_name], contexts[context_name], resolver
                    )
                )
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        workbook.close()
    return report
