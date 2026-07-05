"""Excel-import: tekenconventie, categorie-mapping, placeholders en idempotentie.

Bouwt een mini-werkboek met dezelfde structuur als het echte (blokken onder
elkaar, sectiekoppen met datumrij, Engelse tracking-koppen, Type2-kolom bij
Simon) en importeert het in een geseede in-memory database.
"""

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Budget, Category, Context, Transaction
from app.models.enums import Categorization, CategoryType, TransactionSource
from app.services.budget import build_matrix
from app.services.excel_import import ImportReport, import_workbook


def _build_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()

    budget = wb.active
    assert budget is not None
    budget.title = "Budget Planning"
    # Blok Gemeenschappelijk — labels in kolom B/C, maandkoppen naast de sectiekop
    budget["B2"] = "Budget Planning Gemeenschappelijk"
    budget["C5"] = "Define starting year in settings -->"
    budget["C7"] = "To be allocated:"
    budget["E7"] = 12345.67  # berekende rij: moet genegeerd worden
    budget["C9"] = "Inkomen"
    budget["E9"] = datetime(2025, 1, 1)
    budget["F9"] = datetime(2025, 2, 1)
    budget["C10"] = "Gemeenschappelijke bijdrage"
    budget["E10"] = 3000
    budget["F10"] = 3000
    budget["C11"] = "Enter  Income Category  here"  # placeholder
    budget["E11"] = 999
    budget["C12"] = "Total"  # totaalrij
    budget["E12"] = 3000
    budget["C14"] = "Uitgaven"
    budget["E14"] = datetime(2025, 1, 1)
    budget["F14"] = datetime(2025, 2, 1)
    budget["C15"] = "Boodschappen"
    budget["E15"] = 500
    budget["C16"] = "Ontspanning/Sport/Boeken/Abonnementen"  # langer dan seed-naam
    budget["E16"] = 20
    budget["C18"] = "Sparen"
    budget["E18"] = datetime(2025, 1, 1)
    budget["F18"] = datetime(2025, 2, 1)
    budget["C19"] = "Spaarrekening"
    budget["E19"] = 400
    budget["F19"] = -1000  # geplande opname: negatief budget blijft negatief
    # Blok Simon
    budget["B21"] = "Budget Planning Simon"
    budget["C23"] = "Inkomen2"
    budget["E23"] = datetime(2025, 1, 1)
    budget["C24"] = "Zakgeld potje"  # bestaat niet in de seed → auto-aanmaak
    budget["E24"] = 100

    gem = wb.create_sheet("Tracking Gem.")
    gem["C11"], gem["D11"], gem["E11"], gem["F11"], gem["G11"], gem["H11"], gem["I11"] = (
        "Date", "Type", "Category", "Amount", "Details", "Balance", "Effective Date",
    )
    rows = [
        # december-loon met budgetmaand januari
        (datetime(2024, 12, 26), "Inkomen", "Gemeenschappelijke bijdrage", 3000, None,
         datetime(2025, 1, 1)),
        (datetime(2025, 1, 3), "Uitgaven", "Boodschappen", 7.92, "Pluggen Bol.com", None),
        # negatieve uitgave = terugboeking → wordt positief
        (datetime(2025, 1, 5), "Uitgaven", "Boodschappen", -5, "teruggave", None),
        (datetime(2025, 1, 7), "Sparen", "Spaarrekening", 100, None, None),
    ]
    for i, (d, t, c, a, det, eff) in enumerate(rows, start=12):
        gem.cell(row=i, column=3, value=d)
        gem.cell(row=i, column=4, value=t)
        gem.cell(row=i, column=5, value=c)
        gem.cell(row=i, column=6, value=a)
        gem.cell(row=i, column=7, value=det)
        gem.cell(row=i, column=9, value=eff)
    gem.cell(row=16, column=3, value="geen datum")  # rommelrij → overslaan
    gem.cell(row=16, column=6, value=1)

    simon = wb.create_sheet("Tracking S.")
    for col, header in enumerate(
        ["Date", "Type2", "Type", "Category", "Amount", "Details", "Balance", "Effective Date"],
        start=3,
    ):
        simon.cell(row=11, column=col, value=header)
    simon.cell(row=12, column=3, value=datetime(2025, 1, 2))
    simon.cell(row=12, column=4, value="Inkomen")  # Type2 = het échte type
    simon.cell(row=12, column=5, value="Inkomen2")  # Type = hulpjeslabel
    simon.cell(row=12, column=6, value="Loon")  # niet in de seed → auto-aanmaak
    simon.cell(row=12, column=7, value=1530.66)

    wb.save(path)


@pytest.fixture
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "mini-tracker.xlsx"
    _build_workbook(path)
    return path


@pytest.fixture
def report(seeded_db: Session, workbook_path: Path) -> ImportReport:
    return import_workbook(seeded_db, workbook_path)


def _gem(db: Session) -> Context:
    return db.scalars(select(Context).where(Context.name == "Gemeenschappelijk")).one()


class TestBudgetImport:
    def test_cellen_en_placeholders(self, report: ImportReport) -> None:
        gem_block = next(b for b in report.budget_blocks if b.context == "Gemeenschappelijk")
        # bijdrage jan+feb, boodschappen jan, ontspanning jan, spaar jan+feb = 6 cellen
        assert gem_block.cells_new == 6
        assert gem_block.placeholders_skipped >= 2  # "Enter …" en "Total"
        assert gem_block.years == [2025]

    def test_negatief_spaarbudget_blijft_negatief(
        self, report: ImportReport, seeded_db: Session
    ) -> None:
        ctx = _gem(seeded_db)
        spaar = seeded_db.scalars(
            select(Category).where(Category.context_id == ctx.id, Category.name == "Spaarrekening")
        ).one()
        feb = seeded_db.scalars(
            select(Budget).where(
                Budget.category_id == spaar.id, Budget.year == 2025, Budget.month == 2
            )
        ).one()
        assert feb.amount == Decimal("-1000.00")

    def test_langere_naam_mapt_op_seed_categorie(
        self, report: ImportReport, seeded_db: Session
    ) -> None:
        ctx = _gem(seeded_db)
        namen = set(
            seeded_db.scalars(select(Category.name).where(Category.context_id == ctx.id))
        )
        assert "Ontspanning/Sport/Boeken/Abonnementen" not in namen  # gemapt, niet aangemaakt
        assert any("Ontspanning" in m for m in report.name_mappings)

    def test_onbekende_categorie_aangemaakt_bij_simon(self, report: ImportReport) -> None:
        assert "Simon / Inkomen / Zakgeld potje" in report.categories_created

    def test_tba_uit_geimporteerde_budgetten(
        self, report: ImportReport, seeded_db: Session
    ) -> None:
        matrix = build_matrix(seeded_db, _gem(seeded_db), 2025)
        # jan: 3000 − (500 + 20) − 400 = 2080
        assert matrix.to_be_allocated_cents[0] == 208000
        # feb: 3000 − 0 − (−1000) = 4000
        assert matrix.to_be_allocated_cents[1] == 400000


class TestTrackingImport:
    def test_tekenconventie(self, report: ImportReport, seeded_db: Session) -> None:
        gem_report = next(t for t in report.tracking if t.sheet == "Tracking Gem.")
        assert gem_report.imported == 4
        amounts = {
            (tx.type, str(tx.description or "")): tx.amount
            for tx in seeded_db.scalars(
                select(Transaction).where(Transaction.context_id == _gem(seeded_db).id)
            )
        }
        assert amounts[(CategoryType.INKOMEN, "")] == Decimal("3000.00")
        assert amounts[(CategoryType.UITGAVEN, "Pluggen Bol.com")] == Decimal("-7.92")
        assert amounts[(CategoryType.UITGAVEN, "teruggave")] == Decimal("5.00")  # omgekeerd
        assert amounts[(CategoryType.SPAREN, "")] == Decimal("-100.00")

    def test_effective_date_en_fallback(self, report: ImportReport, seeded_db: Session) -> None:
        loon = seeded_db.scalars(
            select(Transaction).where(Transaction.date == date(2024, 12, 26))
        ).one()
        assert loon.effective_date == date(2025, 1, 1)
        boodschap = seeded_db.scalars(
            select(Transaction).where(Transaction.date == date(2025, 1, 3))
        ).one()
        assert boodschap.effective_date == date(2025, 1, 3)  # fallback = date

    def test_type2_kolom_bij_simon(self, report: ImportReport, seeded_db: Session) -> None:
        simon = seeded_db.scalars(select(Context).where(Context.name == "Simon")).one()
        tx = seeded_db.scalars(
            select(Transaction).where(Transaction.context_id == simon.id)
        ).one()
        assert tx.type == CategoryType.INKOMEN
        assert tx.amount == Decimal("1530.66")
        assert tx.source == TransactionSource.IMPORT_EXCEL
        assert tx.categorization == Categorization.MANUAL

    def test_rommelrij_overgeslagen(self, report: ImportReport) -> None:
        gem_report = next(t for t in report.tracking if t.sheet == "Tracking Gem.")
        assert any("rij 16" in reden for reden in gem_report.skipped)


class TestIdempotentie:
    def test_tweede_import_voegt_niets_toe(
        self, report: ImportReport, seeded_db: Session, workbook_path: Path
    ) -> None:
        tweede = import_workbook(seeded_db, workbook_path)
        assert sum(t.imported for t in tweede.tracking) == 0
        assert sum(t.duplicates for t in tweede.tracking) == 5
        assert all(b.cells_new == 0 and b.cells_updated == 0 for b in tweede.budget_blocks)
        assert tweede.categories_created == []
        assert seeded_db.scalars(select(Transaction)).all()  # data van ronde 1 staat er nog
