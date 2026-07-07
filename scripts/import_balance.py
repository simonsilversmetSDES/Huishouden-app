"""Eenmalige migratie van de vermogenshistoriek (spec §6 + §9) uit het werkboek.

Fase A (--inspect): de tabbladen Rekeningstatus en Status balans verkennen —
gevonden contextblokken, rekeningkolommen (+ voorgestelde mapping) en
activaklasse-rijen — zonder iets te schrijven.
Fase B (import): pas afwerken na akkoord over fase A; draait op een KOPIE van de db.

Gebruik (vanuit de repo-root, met de backend-venv):
    backend/.venv/Scripts/python scripts/import_balance.py --inspect
    backend/.venv/Scripts/python scripts/import_balance.py --db sqlite:///data/db/kopie.db

Leest read-only met gecachte formule-waarden (data_only=True).
"""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

# Windows-console: forceer UTF-8 zodat €, − en … niet crashen
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL_DIR = REPO_ROOT / "data" / "excel"


def find_default_workbook() -> Path:
    candidates = sorted(
        [*DEFAULT_EXCEL_DIR.glob("*.xlsm"), *DEFAULT_EXCEL_DIR.glob("*.xlsx")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        sys.exit(f"Geen Excel-bestand gevonden in {DEFAULT_EXCEL_DIR}")
    return candidates[0]


def _fmt(value: Any) -> str:
    if value is None:
        return "."
    if isinstance(value, datetime | date):
        return value.strftime("%d/%m/%y")
    if isinstance(value, float):
        return f"{value:g}"
    return str(value).replace("\n", " ").strip()[:16]


def inspect_workbook(path: Path) -> None:
    from app.services.excel_balance_import import (
        CONTEXT_NAMES,
        REKENINGSTATUS_SHEET,
        STATUS_BALANS_SHEET,
        _asset_for,
        _context_in_row,
        _norm,
    )

    print(f"Werkboek: {path.name}\n")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        # Rekeningstatus: blokken + rekeningkolommen
        if REKENINGSTATUS_SHEET in wb.sheetnames:
            print(f"=== {REKENINGSTATUS_SHEET} ===")
            rows = list(wb[REKENINGSTATUS_SHEET].iter_rows(values_only=True))
            in_block = False
            for r, row in enumerate(rows, start=1):
                name = _context_in_row(row)
                if name in CONTEXT_NAMES:
                    print(f"  r{r}: blok '{name}'")
                    in_block = True
                    continue
                if in_block and any(
                    isinstance(v, str) and _norm(v) == "datum" for v in row
                ):
                    headers = [v.strip() for v in row if isinstance(v, str) and v.strip()]
                    print(f"    kolommen: {headers}")
                    in_block = False

        # Status balans: blokken + activaklasse-rijen
        if STATUS_BALANS_SHEET in wb.sheetnames:
            print(f"\n=== {STATUS_BALANS_SHEET} ===")
            rows = list(wb[STATUS_BALANS_SHEET].iter_rows(values_only=True))
            for r, row in enumerate(rows, start=1):
                name = _context_in_row(row)
                if name in CONTEXT_NAMES:
                    date_row = rows[r] if r < len(rows) else ()
                    dates = [_fmt(v) for v in date_row if isinstance(v, datetime | date)]
                    span = f"{dates[0]}–{dates[-1]} ({len(dates)} maanden)" if dates else "geen"
                    print(f"  r{r}: blok '{name}', datums {span}")
                elif (
                    label := next((v for v in row if isinstance(v, str) and v.strip()), None)
                ) and _asset_for(label) is not None:
                    print(f"      activaklasse-rij: '{label.strip()}' → {_asset_for(label)}")
    finally:
        wb.close()


def run_import(path: Path, db_url: str) -> None:
    """Fase B: importeren + verificatierapport. Draai op een KOPIE van de db."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session

    from app.services.excel_balance_import import import_balance_workbook

    engine = create_engine(db_url)
    with Session(engine) as db:
        report = import_balance_workbook(db, path)

    print(f"Werkboek: {path.name}")
    print(f"Database: {db_url}\n")

    print("— Rekeningstatus (§6) —")
    for account_block in report.accounts:
        print(
            f"  {account_block.context:<18} {account_block.snapshots_new:>4} nieuw, "
            f"{account_block.snapshots_updated} bijgewerkt"
        )
        for mapped in account_block.mapped_columns:
            print(f"      ✓ {mapped}")
        for unmapped in account_block.unmapped_columns:
            print(f"      · overgeslagen kolom: {unmapped}")

    print("\n— Status balans (§9) —")
    for nw_block in report.net_worth:
        print(
            f"  {nw_block.context:<18} {nw_block.values_new:>4} nieuw, "
            f"{nw_block.values_updated} bijgewerkt"
        )
        for unmapped in nw_block.unmapped_rows:
            print(f"      · overgeslagen rij: {unmapped}")

    # Hercheck: Gemeenschappelijk 01/02/2025 rekeningtotaal moet € 3.143,00 zijn.
    _verify(db_url)


def _verify(db_url: str) -> None:
    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import Session

    from app.models import Context
    from app.services.account_status import build_account_status

    engine = create_engine(db_url)
    with Session(engine) as db:
        gem = db.scalars(select(Context).where(Context.name == "Gemeenschappelijk")).one()
        status = build_account_status(db, gem, today=date(2026, 1, 1))
        feb = next((r for r in status.rows if r.snapshot_date == date(2025, 2, 1)), None)
        got = feb.total_cents if feb else None
        verdict = "OK" if got == 314300 else "WIJKT AF!"
        print(
            f"\nHercheck rekeningtotaal Gem. 01/02/2025: "
            f"€ {(got or 0) / 100:.2f} (verwacht € 3.143,00) → {verdict}"
        )
        if got != 314300:
            sys.exit(1)


def main() -> None:
    parser = argparse.ArgumentParser(description="Balans-migratie voor de Huishouden-app")
    parser.add_argument("--file", type=Path, default=None, help="pad naar het .xlsm-werkboek")
    parser.add_argument("--inspect", action="store_true", help="alleen verkennen, niets schrijven")
    parser.add_argument(
        "--db",
        default=None,
        help=(
            "SQLAlchemy-URL van de DOELDATABASE (verplicht voor import). Gebruik een KOPIE, "
            "nooit de echte database — bv. sqlite:///data/db/kopie.db"
        ),
    )
    args = parser.parse_args()

    sys.path.insert(0, str(REPO_ROOT / "backend"))

    path = args.file or find_default_workbook()
    if not path.exists():
        sys.exit(f"Bestand niet gevonden: {path}")

    if args.inspect:
        inspect_workbook(path)
        return
    if not args.db:
        sys.exit(
            "Import vereist --db met de URL van een KOPIE van de database "
            "(bv. --db sqlite:///data/db/kopie.db). Zonder --db gebeurt er niets."
        )
    run_import(path, args.db)


if __name__ == "__main__":
    main()
