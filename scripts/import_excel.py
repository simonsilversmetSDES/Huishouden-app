"""Eenmalige Excel-import: het bestaande Budget Tracker-werkboek -> de app-database.

Fase A (--inspect): het werkboek verkennen zonder iets te schrijven — tabbladen,
gevonden jaren, kolomstructuur en de tekenconventie (+/-) van de bedragen.
Fase B (import): wordt pas afgewerkt na akkoord over de bevindingen van fase A.

Gebruik (vanuit de repo-root, met de backend-venv):
    backend/.venv/Scripts/python scripts/import_excel.py --inspect
    backend/.venv/Scripts/python scripts/import_excel.py --inspect --file "data/excel/....xlsm"

Leest het werkboek altijd read-only met de door Excel gecachte formule-waarden
(data_only=True); het bronbestand wordt nooit gewijzigd.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Windows-console gebruikt cp1252; forceer UTF-8 zodat €, − en … niet crashen
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL_DIR = REPO_ROOT / "data" / "excel"

# Tabbladen waar we dieper in graven tijdens --inspect
BUDGET_KEYWORDS = ("budget", "planning")
TRACKING_KEYWORDS = ("tracking",)

YEAR_RE = re.compile(r"\b(20[1-3][0-9])\b")
MAAND_NAMEN = (
    "jan",
    "feb",
    "maa",
    "apr",
    "mei",
    "jun",
    "jul",
    "aug",
    "sep",
    "okt",
    "nov",
    "dec",
)


def find_default_workbook() -> Path:
    """Pak het (nieuwste) .xlsm/.xlsx-bestand uit data/excel/."""
    candidates = sorted(
        [*DEFAULT_EXCEL_DIR.glob("*.xlsm"), *DEFAULT_EXCEL_DIR.glob("*.xlsx")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        sys.exit(f"Geen Excel-bestand gevonden in {DEFAULT_EXCEL_DIR}")
    return candidates[0]


def fmt_cell(value: Any) -> str:
    """Compacte weergave van een celwaarde voor de preview-grid."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    text = str(value).replace("\n", " ").strip()
    return text[:18] + "…" if len(text) > 19 else text


def sheet_rows(ws: Worksheet, max_rows: int | None = None) -> list[tuple[Any, ...]]:
    rows: list[tuple[Any, ...]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows is not None and i >= max_rows:
            break
        rows.append(row)
    return rows


def print_preview(
    rows: list[tuple[Any, ...]], n_rows: int = 14, n_cols: int = 15
) -> None:
    """Print de linkerbovenhoek van een tabblad als grid (rij- en kolomnummers 1-based)."""
    header = "      " + " ".join(f"[{c:>2}]" + " " * 15 for c in range(1, n_cols + 1))
    print(header.rstrip())
    for r, row in enumerate(rows[:n_rows], start=1):
        cells = [fmt_cell(v).ljust(19) for v in row[:n_cols]]
        print(f"r{r:>3} | " + " ".join(cells).rstrip())


def detect_years(name: str, rows: list[tuple[Any, ...]]) -> list[int]:
    """Jaartallen in de tabbladnaam en in de eerste rijen van het blad."""
    found: set[int] = {int(m) for m in YEAR_RE.findall(name)}
    for row in rows[:20]:
        for value in row:
            if (
                isinstance(value, int | float)
                and 2010 <= value <= 2039
                and value == int(value)
            ):
                found.add(int(value))
            elif isinstance(value, str):
                found.update(int(m) for m in YEAR_RE.findall(value))
            elif isinstance(value, datetime | date):
                found.add(value.year)
    return sorted(found)


def find_header_row(rows: list[tuple[Any, ...]], *needles: str) -> int | None:
    """Zoek (0-based) de rij waarin alle gegeven kolomkoppen voorkomen."""
    for i, row in enumerate(rows[:30]):
        texts = {str(v).strip().lower() for v in row if isinstance(v, str)}
        if all(any(needle in t for t in texts) for needle in needles):
            return i
    return None


def date_row_summary(rows: list[tuple[Any, ...]]) -> None:
    """Rijen met veel datumcellen = maand-kolomkoppen; toon bereik per rij."""
    for r, row in enumerate(rows, start=1):
        dates = [v for v in row if isinstance(v, datetime | date)]
        if len(dates) >= 6:
            print(
                f"  Maandkoppen r{r:>3}: {len(dates)} datums, "
                f"{fmt_cell(dates[0])} t/m {fmt_cell(dates[-1])}"
            )


def analyse_signs(
    rows: list[tuple[Any, ...]], header_idx: int, headers: list[str]
) -> None:
    """Tekenconventie: verdeling +/−/0 van de bedragen, gegroepeerd per type-kolom."""

    def col(*needles: str) -> int | None:
        for j, h in enumerate(headers):
            if any(n in h for n in needles):
                return j
        return None

    amount_col = col("bedrag", "amount")
    type_col = col("type", "soort")
    date_col = col("datum", "date")
    if amount_col is None:
        print("  !! Geen bedrag-kolom herkend — tekenanalyse overgeslagen")
        return

    per_type: dict[str, Counter[str]] = {}
    per_year: Counter[int] = Counter()
    samples: dict[str, list[str]] = {}
    for row in rows[header_idx + 1 :]:
        if amount_col >= len(row):
            continue
        amount = row[amount_col]
        if not isinstance(amount, int | float | Decimal) or isinstance(amount, bool):
            continue
        type_value = "(geen type-kolom)"
        if type_col is not None and type_col < len(row) and row[type_col] is not None:
            type_value = str(row[type_col]).strip()
        sign = "+" if amount > 0 else ("−" if amount < 0 else "0")
        per_type.setdefault(type_value, Counter())[sign] += 1
        if len(samples.setdefault(type_value, [])) < 3:
            desc = ""
            if date_col is not None and date_col < len(row):
                desc = fmt_cell(row[date_col])
            samples[type_value].append(f"{desc} {amount:,.2f}".strip())
        if date_col is not None and date_col < len(row):
            d = row[date_col]
            if isinstance(d, datetime | date):
                per_year[d.year] += 1

    print("  Tekenconventie per type (aantal +/−/0):")
    for type_value, counts in sorted(per_type.items()):
        total = sum(counts.values())
        print(
            f"    {type_value:<28} n={total:<5} "
            f"+:{counts['+']:<5} −:{counts['−']:<5} 0:{counts['0']:<4} "
            f"vb: {'; '.join(samples.get(type_value, []))}"
        )
    if per_year:
        years = ", ".join(f"{y}: {n}" for y, n in sorted(per_year.items()))
        print(f"  Transacties per jaar (uit de datum-kolom): {years}")


def inspect_budget_sheet(ws: Worksheet) -> None:
    rows = sheet_rows(ws)
    print_preview(rows)
    years = detect_years(ws.title, rows)
    print(f"  Gevonden jaartallen: {years or 'geen'}")

    # Waar beginnen de context-blokken (Gem./Simon/Jozefien) en de sectiekoppen?
    for r, row in enumerate(rows, start=1):
        for value in row[:4]:
            if isinstance(value, str):
                text = value.strip()
                if text.startswith("Budget Planning") or text.lower().startswith(
                    (
                        "inkomen",
                        "uitgaven",
                        "sparen",
                        "to be allocated",
                        "define starting",
                    )
                ):
                    print(f"  Structuur r{r:>3}: {text}")
                    break

    date_row_summary(rows)

    # Negatieve budgetcellen verraden de tekenconventie; de TBA-rijen zijn
    # berekende saldi en horen niet bij de invoer — apart tellen.
    tba_rows = {
        r
        for r, row in enumerate(rows, start=1)
        for v in row[:4]
        if isinstance(v, str) and "to be allocated" in v.lower()
    }
    neg = pos = tba_neg = 0
    neg_samples: list[str] = []
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            if (
                isinstance(value, int | float)
                and not isinstance(value, bool)
                and value != 0
            ):
                if value < 0:
                    if r in tba_rows:
                        tba_neg += 1
                    else:
                        neg += 1
                        if len(neg_samples) < 8:
                            neg_samples.append(f"r{r}c{c}={value:,.2f}")
                else:
                    pos += 1
    print(
        f"  Numerieke cellen: {pos} positief, {neg} negatief buiten de TBA-rijen "
        f"({tba_neg} negatief in TBA-rijen {sorted(tba_rows)})"
    )
    print(f"  Negatieve cellen buiten TBA: {neg_samples or 'geen'}")


def inspect_tracking_sheet(ws: Worksheet) -> None:
    rows = sheet_rows(ws)
    print_preview(rows)
    years = detect_years(ws.title, rows)
    print(f"  Gevonden jaartallen: {years or 'geen'}")

    # Koppen zijn Engels in dit werkboek ("Date", "Amount"); Nederlands als fallback
    header_idx = find_header_row(rows, "date", "amount")
    if header_idx is None:
        header_idx = find_header_row(rows, "datum", "bedrag")
    if header_idx is None:
        print(
            "  !! Geen kop-rij met Date/Amount of Datum/Bedrag gevonden in de eerste 30 rijen"
        )
        return
    headers = [
        str(v).strip().lower() if v is not None else "" for v in rows[header_idx]
    ]
    print(f"  Kop-rij r{header_idx + 1}: {[h for h in headers if h]}")
    analyse_signs(rows, header_idx, headers)


def inspect_workbook(path: Path) -> None:
    print(f"Werkboek: {path.name}\n")
    wb: Workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        print(f"{'Tabblad':<32} {'rijen':>6} {'kolommen':>9}")
        print("-" * 49)
        for ws in wb.worksheets:
            print(f"{ws.title:<32} {ws.max_row or 0:>6} {ws.max_column or 0:>9}")

        for ws in wb.worksheets:
            name = ws.title.lower()
            if any(k in name for k in BUDGET_KEYWORDS):
                print(f"\n=== BUDGET-TABBLAD: {ws.title} " + "=" * 30)
                inspect_budget_sheet(ws)
            elif any(k in name for k in TRACKING_KEYWORDS):
                print(f"\n=== TRACKING-TABBLAD: {ws.title} " + "=" * 30)
                inspect_tracking_sheet(ws)
    finally:
        wb.close()


def run_import(path: Path, db_url: str) -> None:
    """Fase B: importeren + verificatierapport. Draai dit op een KOPIE van de db."""
    sys.path.insert(0, str(REPO_ROOT / "backend"))
    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import Session

    from app.models import Context
    from app.services.budget import build_matrix
    from app.services.excel_import import import_workbook

    engine = create_engine(db_url)
    with Session(engine) as db:
        report = import_workbook(db, path)

        print(f"Werkboek: {path.name}")
        print(f"Database: {db_url}\n")

        print("— Budget Planning —")
        for block in report.budget_blocks:
            years = f"{block.years[0]}–{block.years[-1]}" if block.years else "geen"
            print(
                f"  {block.context:<18} {block.cells_new:>5} nieuwe cellen, "
                f"{block.cells_updated} bijgewerkt, {block.placeholders_skipped} "
                f"placeholder/totaal-rijen overgeslagen (jaren: {years})"
            )

        print("\n— Tracking —")
        for tracking in report.tracking:
            print(
                f"  {tracking.sheet:<16} ({tracking.context}): "
                f"{tracking.imported} geïmporteerd, {tracking.duplicates} duplicaten, "
                f"{len(tracking.skipped)} overgeslagen"
            )
            for reden in tracking.skipped:
                print(f"      · {reden}")

        if report.name_mappings:
            print("\n— Naam-mappings (Excel → bestaande categorie) —")
            for mapping in report.name_mappings:
                print(f"  {mapping}")
        if report.categories_created:
            print("\n— Automatisch aangemaakte categorieën —")
            for created in report.categories_created:
                print(f"  {created}")
        else:
            print("\nGeen categorieën automatisch aangemaakt.")

        # Hercheck spec §10: TBA jan 2025 (Gemeenschappelijk) moet € 92,08 zijn.
        gem = db.scalars(select(Context).where(Context.name == "Gemeenschappelijk")).one()
        tba_cents = build_matrix(db, gem, 2025).to_be_allocated_cents[0]
        verdict = "OK" if tba_cents == 9208 else "WIJKT AF!"
        print(
            f"\nTBA-hercheck jan 2025 (Gem.): € {tba_cents / 100:.2f} "
            f"(verwacht € 92,08) → {verdict}"
        )
        if tba_cents != 9208:
            sys.exit(1)


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel-import voor de Huishouden-app")
    parser.add_argument(
        "--file", type=Path, default=None, help="pad naar het .xlsm-werkboek"
    )
    parser.add_argument(
        "--inspect", action="store_true", help="alleen verkennen, niets schrijven"
    )
    parser.add_argument(
        "--db",
        default=None,
        help=(
            "SQLAlchemy-URL van de DOELDATABASE (verplicht voor import). "
            "Gebruik een kopie, nooit de echte database — bv. "
            "sqlite:///data/db/import-kopie.db"
        ),
    )
    args = parser.parse_args()

    path = args.file or find_default_workbook()
    if not path.exists():
        sys.exit(f"Bestand niet gevonden: {path}")

    if args.inspect:
        inspect_workbook(path)
        return

    if not args.db:
        sys.exit(
            "Import vereist --db met de URL van een KOPIE van de database "
            "(bv. --db sqlite:///data/db/import-kopie.db). Zonder --db gebeurt er niets."
        )
    run_import(path, args.db)


if __name__ == "__main__":
    main()
